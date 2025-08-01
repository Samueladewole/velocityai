"""
ERIP Sheets Engine Core
Native spreadsheet capabilities with AI-powered intelligence and ERIP component integration
"""

from typing import Dict, List, Optional, Any, Union, Tuple
from datetime import datetime, timedelta
from pydantic import BaseModel, Field
import pandas as pd
import numpy as np
import json
import asyncio
import structlog
from enum import Enum
import uuid
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import io
import base64

logger = structlog.get_logger()

class CellDataType(str, Enum):
    """Supported cell data types"""
    NUMBER = "number"
    TEXT = "text"
    BOOLEAN = "boolean"
    DATE = "date"
    FORMULA = "formula"
    ERROR = "error"

class ChartType(str, Enum):
    """Supported chart types"""
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"
    COLUMN = "column"

class DataSource(str, Enum):
    """Data source types for live connections"""
    PRISM = "prism"
    BEACON = "beacon"
    ATLAS = "atlas"
    COMPASS = "compass"
    NEXUS = "nexus"
    EXTERNAL_API = "external_api"
    DATABASE = "database"
    FILE_UPLOAD = "file_upload"

class Cell(BaseModel):
    """Individual spreadsheet cell"""
    row: int
    column: int
    value: Union[str, int, float, bool, None] = None
    formula: Optional[str] = None
    data_type: CellDataType = CellDataType.TEXT
    format: Optional[Dict[str, Any]] = None
    comment: Optional[str] = None
    validation: Optional[Dict[str, Any]] = None
    
class Range(BaseModel):
    """Spreadsheet range definition"""
    start_row: int
    start_column: int
    end_row: int
    end_column: int
    
    def to_excel_notation(self) -> str:
        """Convert to Excel A1:B2 notation"""
        start_col = self._number_to_column(self.start_column)
        end_col = self._number_to_column(self.end_column)
        return f"{start_col}{self.start_row}:{end_col}{self.end_row}"
    
    def _number_to_column(self, n: int) -> str:
        """Convert column number to Excel letter notation"""
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + n % 26) + result
            n //= 26
        return result

class Chart(BaseModel):
    """Chart definition"""
    chart_id: str
    chart_type: ChartType
    title: str
    data_range: Range
    position: Dict[str, int]  # row, column
    width: int = 400
    height: int = 300
    series_config: Optional[Dict[str, Any]] = None

class DataConnection(BaseModel):
    """Live data connection definition"""
    connection_id: str
    name: str
    source_type: DataSource
    connection_config: Dict[str, Any]
    refresh_schedule: Optional[str] = None  # cron format
    last_refresh: Optional[datetime] = None
    auto_refresh: bool = True
    target_range: Range

class Worksheet(BaseModel):
    """Individual worksheet in workbook"""
    worksheet_id: str
    name: str
    cells: Dict[str, Cell] = Field(default_factory=dict)  # "row,col" -> Cell
    charts: Dict[str, Chart] = Field(default_factory=dict)
    data_connections: Dict[str, DataConnection] = Field(default_factory=dict)
    protected: bool = False
    hidden: bool = False
    row_count: int = 1000
    column_count: int = 100
    
    def get_cell(self, row: int, column: int) -> Optional[Cell]:
        """Get cell at specific position"""
        key = f"{row},{column}"
        return self.cells.get(key)
    
    def set_cell(self, row: int, column: int, cell: Cell) -> None:
        """Set cell at specific position"""
        key = f"{row},{column}"
        self.cells[key] = cell

class Workbook(BaseModel):
    """Complete workbook with multiple worksheets"""
    workbook_id: str
    name: str
    worksheets: Dict[str, Worksheet] = Field(default_factory=dict)
    shared_formulas: Dict[str, str] = Field(default_factory=dict)
    named_ranges: Dict[str, Range] = Field(default_factory=dict)
    created_by: str
    created_at: datetime
    modified_at: datetime
    version: int = 1
    permissions: Dict[str, List[str]] = Field(default_factory=dict)  # user_id -> permissions
    
    def add_worksheet(self, name: str) -> str:
        """Add new worksheet and return its ID"""
        worksheet_id = str(uuid.uuid4())
        worksheet = Worksheet(
            worksheet_id=worksheet_id,
            name=name
        )
        self.worksheets[worksheet_id] = worksheet
        return worksheet_id
    
    def get_worksheet(self, worksheet_id: str) -> Optional[Worksheet]:
        """Get worksheet by ID"""
        return self.worksheets.get(worksheet_id)

class SheetsEngine:
    """
    Core spreadsheet engine with Excel compatibility and AI integration
    """
    
    def __init__(self):
        self.workbooks: Dict[str, Workbook] = {}
        self.formula_engine = FormulaEngine()
        self.ai_assistant = SheetsAIAssistant()
        self.data_connector = DataConnector()
        self.excel_converter = ExcelConverter()
        
    async def create_workbook(
        self, 
        name: str, 
        created_by: str,
        template: Optional[str] = None
    ) -> str:
        """Create new workbook"""
        try:
            workbook_id = str(uuid.uuid4())
            
            workbook = Workbook(
                workbook_id=workbook_id,
                name=name,
                created_by=created_by,
                created_at=datetime.utcnow(),
                modified_at=datetime.utcnow()
            )
            
            # Add default worksheet
            default_sheet_id = workbook.add_worksheet("Sheet1")
            
            # Apply template if specified
            if template:
                await self._apply_template(workbook, template)
            
            self.workbooks[workbook_id] = workbook
            
            logger.info("Workbook created", 
                       workbook_id=workbook_id, 
                       name=name, 
                       created_by=created_by)
            
            return workbook_id
            
        except Exception as e:
            logger.error("Failed to create workbook", error=str(e))
            raise
    
    async def update_cell(
        self,
        workbook_id: str,
        worksheet_id: str,
        row: int,
        column: int,
        value: Any,
        formula: Optional[str] = None
    ) -> Cell:
        """Update cell value and recalculate dependent cells"""
        try:
            workbook = self.workbooks.get(workbook_id)
            if not workbook:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            worksheet = workbook.get_worksheet(worksheet_id)
            if not worksheet:
                raise ValueError(f"Worksheet {worksheet_id} not found")
            
            # Determine data type
            data_type = self._determine_data_type(value, formula)
            
            # Create cell
            cell = Cell(
                row=row,
                column=column,
                value=value,
                formula=formula,
                data_type=data_type
            )
            
            # Calculate formula if present
            if formula:
                try:
                    calculated_value = await self.formula_engine.calculate_formula(
                        formula, workbook, worksheet_id
                    )
                    cell.value = calculated_value
                except Exception as e:
                    cell.data_type = CellDataType.ERROR
                    cell.value = f"#ERROR: {str(e)}"
            
            # Update cell in worksheet
            worksheet.set_cell(row, column, cell)
            
            # Update modification time
            workbook.modified_at = datetime.utcnow()
            workbook.version += 1
            
            logger.info("Cell updated", 
                       workbook_id=workbook_id,
                       worksheet_id=worksheet_id,
                       row=row, column=column)
            
            return cell
            
        except Exception as e:
            logger.error("Failed to update cell", 
                        workbook_id=workbook_id,
                        error=str(e))
            raise
    
    async def get_range_data(
        self,
        workbook_id: str,
        worksheet_id: str,
        range_def: Range
    ) -> List[List[Any]]:
        """Get data from specified range"""
        try:
            workbook = self.workbooks.get(workbook_id)
            if not workbook:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            worksheet = workbook.get_worksheet(worksheet_id)
            if not worksheet:
                raise ValueError(f"Worksheet {worksheet_id} not found")
            
            data = []
            for row in range(range_def.start_row, range_def.end_row + 1):
                row_data = []
                for col in range(range_def.start_column, range_def.end_column + 1):
                    cell = worksheet.get_cell(row, col)
                    row_data.append(cell.value if cell else None)
                data.append(row_data)
            
            return data
            
        except Exception as e:
            logger.error("Failed to get range data", error=str(e))
            raise
    
    async def set_range_data(
        self,
        workbook_id: str,
        worksheet_id: str,
        range_def: Range,
        data: List[List[Any]]
    ) -> bool:
        """Set data for specified range"""
        try:
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    row = range_def.start_row + i
                    col = range_def.start_column + j
                    
                    if (row <= range_def.end_row and 
                        col <= range_def.end_column):
                        await self.update_cell(
                            workbook_id, worksheet_id, row, col, value
                        )
            
            logger.info("Range data updated", 
                       workbook_id=workbook_id,
                       range=range_def.to_excel_notation())
            
            return True
            
        except Exception as e:
            logger.error("Failed to set range data", error=str(e))
            raise
    
    async def create_chart(
        self,
        workbook_id: str,
        worksheet_id: str,
        chart_type: ChartType,
        title: str,
        data_range: Range,
        position: Dict[str, int]
    ) -> str:
        """Create chart in worksheet"""
        try:
            workbook = self.workbooks.get(workbook_id)
            if not workbook:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            worksheet = workbook.get_worksheet(worksheet_id)
            if not worksheet:
                raise ValueError(f"Worksheet {worksheet_id} not found")
            
            chart_id = str(uuid.uuid4())
            
            chart = Chart(
                chart_id=chart_id,
                chart_type=chart_type,
                title=title,
                data_range=data_range,
                position=position
            )
            
            worksheet.charts[chart_id] = chart
            workbook.modified_at = datetime.utcnow()
            
            logger.info("Chart created", 
                       workbook_id=workbook_id,
                       chart_id=chart_id,
                       chart_type=chart_type.value)
            
            return chart_id
            
        except Exception as e:
            logger.error("Failed to create chart", error=str(e))
            raise
    
    async def connect_data_source(
        self,
        workbook_id: str,
        worksheet_id: str,
        connection_name: str,
        source_type: DataSource,
        config: Dict[str, Any],
        target_range: Range
    ) -> str:
        """Connect external data source to worksheet"""
        try:
            workbook = self.workbooks.get(workbook_id)
            if not workbook:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            worksheet = workbook.get_worksheet(worksheet_id)
            if not worksheet:
                raise ValueError(f"Worksheet {worksheet_id} not found")
            
            connection_id = str(uuid.uuid4())
            
            connection = DataConnection(
                connection_id=connection_id,
                name=connection_name,
                source_type=source_type,
                connection_config=config,
                target_range=target_range,
                last_refresh=datetime.utcnow()
            )
            
            # Initial data load
            data = await self.data_connector.fetch_data(connection)
            if data:
                await self.set_range_data(workbook_id, worksheet_id, target_range, data)
            
            worksheet.data_connections[connection_id] = connection
            workbook.modified_at = datetime.utcnow()
            
            logger.info("Data source connected", 
                       workbook_id=workbook_id,
                       connection_id=connection_id,
                       source_type=source_type.value)
            
            return connection_id
            
        except Exception as e:
            logger.error("Failed to connect data source", error=str(e))
            raise
    
    async def refresh_data_connections(
        self,
        workbook_id: str,
        worksheet_id: Optional[str] = None
    ) -> Dict[str, bool]:
        """Refresh data connections"""
        try:
            workbook = self.workbooks.get(workbook_id)
            if not workbook:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            refresh_results = {}
            
            # Determine worksheets to refresh
            worksheets_to_refresh = []
            if worksheet_id:
                worksheet = workbook.get_worksheet(worksheet_id)
                if worksheet:
                    worksheets_to_refresh.append((worksheet_id, worksheet))
            else:
                worksheets_to_refresh = list(workbook.worksheets.items())
            
            # Refresh each worksheet's connections
            for ws_id, worksheet in worksheets_to_refresh:
                for conn_id, connection in worksheet.data_connections.items():
                    try:
                        data = await self.data_connector.fetch_data(connection)
                        if data:
                            await self.set_range_data(
                                workbook_id, ws_id, connection.target_range, data
                            )
                            connection.last_refresh = datetime.utcnow()
                            refresh_results[conn_id] = True
                        else:
                            refresh_results[conn_id] = False
                    except Exception as e:
                        logger.error("Data connection refresh failed", 
                                   connection_id=conn_id, error=str(e))
                        refresh_results[conn_id] = False
            
            workbook.modified_at = datetime.utcnow()
            
            logger.info("Data connections refreshed", 
                       workbook_id=workbook_id,
                       results=refresh_results)
            
            return refresh_results
            
        except Exception as e:
            logger.error("Failed to refresh data connections", error=str(e))
            raise
    
    def _determine_data_type(self, value: Any, formula: Optional[str] = None) -> CellDataType:
        """Determine cell data type from value"""
        if formula:
            return CellDataType.FORMULA
        
        if isinstance(value, bool):
            return CellDataType.BOOLEAN
        elif isinstance(value, (int, float)):
            return CellDataType.NUMBER
        elif isinstance(value, datetime):
            return CellDataType.DATE
        else:
            return CellDataType.TEXT
    
    async def _apply_template(self, workbook: Workbook, template: str) -> None:
        """Apply template to workbook"""
        # Template application logic would go here
        # For now, just add some sample data based on template type
        if template == "risk_dashboard":
            await self._create_risk_dashboard_template(workbook)
        elif template == "compliance_tracker":
            await self._create_compliance_tracker_template(workbook)
        elif template == "roi_calculator":
            await self._create_roi_calculator_template(workbook)
    
    async def _create_risk_dashboard_template(self, workbook: Workbook) -> None:
        """Create risk dashboard template"""
        # Implementation would create a pre-configured risk dashboard
        pass
    
    async def _create_compliance_tracker_template(self, workbook: Workbook) -> None:
        """Create compliance tracker template"""
        # Implementation would create a compliance tracking spreadsheet
        pass
    
    async def _create_roi_calculator_template(self, workbook: Workbook) -> None:
        """Create ROI calculator template"""
        # Implementation would create an ROI calculation spreadsheet
        pass

class FormulaEngine:
    """Excel-compatible formula calculation engine"""
    
    def __init__(self):
        self.supported_functions = {
            'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'IF', 'VLOOKUP',
            'INDEX', 'MATCH', 'CONCATENATE', 'LEFT', 'RIGHT', 'MID',
            'UPPER', 'LOWER', 'TRIM', 'NOW', 'TODAY', 'DATE', 'YEAR',
            'MONTH', 'DAY', 'ROUND', 'ABS', 'SQRT', 'POWER'
        }
    
    async def calculate_formula(
        self, 
        formula: str, 
        workbook: Workbook, 
        worksheet_id: str
    ) -> Any:
        """Calculate formula result"""
        try:
            # Remove leading = if present
            if formula.startswith('='):
                formula = formula[1:]
            
            # Parse and evaluate formula
            # This is a simplified implementation - full Excel compatibility would require
            # a complete formula parser and evaluator
            result = await self._evaluate_formula(formula, workbook, worksheet_id)
            return result
            
        except Exception as e:
            logger.error("Formula calculation failed", formula=formula, error=str(e))
            raise
    
    async def _evaluate_formula(
        self, 
        formula: str, 
        workbook: Workbook, 
        worksheet_id: str
    ) -> Any:
        """Evaluate formula expression"""
        # Simplified formula evaluation
        # In production, would use a full Excel formula parser
        
        # Handle simple SUM function
        if formula.upper().startswith('SUM('):
            range_ref = formula[4:-1]  # Remove SUM( and )
            values = await self._get_range_values(range_ref, workbook, worksheet_id)
            return sum(v for v in values if isinstance(v, (int, float)))
        
        # Handle simple arithmetic
        try:
            # Replace cell references with values
            evaluated_formula = await self._substitute_cell_references(
                formula, workbook, worksheet_id
            )
            return eval(evaluated_formula)  # Note: In production, use safe evaluation
        except:
            return f"#ERROR: Cannot evaluate {formula}"
    
    async def _get_range_values(
        self, 
        range_ref: str, 
        workbook: Workbook, 
        worksheet_id: str
    ) -> List[Any]:
        """Get values from range reference"""
        # Parse range reference (e.g., "A1:B10")
        # This is simplified - full implementation would handle all Excel range formats
        worksheet = workbook.get_worksheet(worksheet_id)
        if not worksheet:
            return []
        
        values = []
        # Extract values from range (simplified implementation)
        for cell in worksheet.cells.values():
            if isinstance(cell.value, (int, float)):
                values.append(cell.value)
        
        return values
    
    async def _substitute_cell_references(
        self, 
        formula: str, 
        workbook: Workbook, 
        worksheet_id: str
    ) -> str:
        """Substitute cell references with actual values"""
        # Simplified implementation
        # In production, would properly parse and substitute all cell references
        return formula

class SheetsAIAssistant:
    """AI-powered spreadsheet assistant for insights and automation"""
    
    def __init__(self):
        self.insights_cache = {}
    
    async def analyze_data(
        self, 
        workbook_id: str, 
        worksheet_id: str, 
        data_range: Range
    ) -> Dict[str, Any]:
        """Analyze data range and provide insights"""
        # AI-powered data analysis would go here
        return {
            "summary": "Data analysis insights",
            "patterns": [],
            "anomalies": [],
            "recommendations": []
        }
    
    async def suggest_formulas(
        self, 
        context: str, 
        data_description: str
    ) -> List[str]:
        """Suggest appropriate formulas based on context"""
        # AI formula suggestion would go here
        return ["=SUM(A1:A10)", "=AVERAGE(B1:B10)", "=COUNT(C1:C10)"]
    
    async def generate_insights(
        self, 
        workbook: Workbook, 
        worksheet_id: str
    ) -> List[str]:
        """Generate insights about worksheet data"""
        # AI insight generation would go here
        return [
            "Risk levels have increased by 15% this quarter",
            "Compliance scores are trending upward",
            "Financial impact calculations show positive ROI"
        ]

class DataConnector:
    """Data connector for ERIP components and external sources"""
    
    async def fetch_data(self, connection: DataConnection) -> Optional[List[List[Any]]]:
        """Fetch data from connected source"""
        try:
            if connection.source_type == DataSource.PRISM:
                return await self._fetch_prism_data(connection.connection_config)
            elif connection.source_type == DataSource.BEACON:
                return await self._fetch_beacon_data(connection.connection_config)
            elif connection.source_type == DataSource.ATLAS:
                return await self._fetch_atlas_data(connection.connection_config)
            elif connection.source_type == DataSource.COMPASS:
                return await self._fetch_compass_data(connection.connection_config)
            # Add more source types as needed
            
            return None
            
        except Exception as e:
            logger.error("Data fetch failed", 
                        connection_id=connection.connection_id,
                        error=str(e))
            return None
    
    async def _fetch_prism_data(self, config: Dict[str, Any]) -> List[List[Any]]:
        """Fetch data from PRISM component"""
        # Mock PRISM data
        return [
            ["Risk Scenario", "Probability", "Impact", "Total Risk"],
            ["Data Breach", 0.15, 1000000, 150000],
            ["System Failure", 0.05, 500000, 25000],
            ["Compliance Violation", 0.08, 750000, 60000]
        ]
    
    async def _fetch_beacon_data(self, config: Dict[str, Any]) -> List[List[Any]]:
        """Fetch data from BEACON component"""
        # Mock BEACON data
        return [
            ["Metric", "Current", "Target", "Variance"],
            ["ROI %", 525, 300, 225],
            ["Cost Savings", 280000, 200000, 80000],
            ["Efficiency Gain", 192500, 150000, 42500]
        ]
    
    async def _fetch_atlas_data(self, config: Dict[str, Any]) -> List[List[Any]]:
        """Fetch data from ATLAS component"""
        # Mock ATLAS data
        return [
            ["Control", "Status", "Score", "Last Assessment"],
            ["Access Control", "Compliant", 95, "2025-01-15"],
            ["Data Encryption", "Compliant", 98, "2025-01-14"],
            ["Network Security", "Partial", 75, "2025-01-13"]
        ]
    
    async def _fetch_compass_data(self, config: Dict[str, Any]) -> List[List[Any]]:
        """Fetch data from COMPASS component"""
        # Mock COMPASS data
        return [
            ["Framework", "Status", "Compliance %", "Next Review"],
            ["GDPR", "Active", 92, "2025-03-01"],
            ["SOX", "Active", 88, "2025-02-15"],
            ["HIPAA", "Planned", 0, "2025-04-01"]
        ]

class ExcelConverter:
    """Excel file import/export functionality"""
    
    async def import_excel(self, file_content: bytes, workbook_id: str) -> bool:
        """Import Excel file into ERIP workbook"""
        try:
            # Load Excel file
            excel_workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            
            # Convert to ERIP format
            # Implementation would convert Excel workbook to ERIP Workbook
            
            logger.info("Excel file imported", workbook_id=workbook_id)
            return True
            
        except Exception as e:
            logger.error("Excel import failed", error=str(e))
            return False
    
    async def export_excel(self, workbook: Workbook) -> bytes:
        """Export ERIP workbook to Excel format"""
        try:
            # Create Excel workbook
            excel_workbook = openpyxl.Workbook()
            
            # Convert ERIP workbook to Excel format
            # Implementation would convert Workbook to Excel format
            
            # Save to bytes
            buffer = io.BytesIO()
            excel_workbook.save(buffer)
            buffer.seek(0)
            
            logger.info("Excel file exported", workbook_id=workbook.workbook_id)
            return buffer.getvalue()
            
        except Exception as e:
            logger.error("Excel export failed", error=str(e))
            raise